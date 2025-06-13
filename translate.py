import os
import google.generativeai as genai
import openpyxl
import time

# Initialize Gemini API
API_KEY = os.getenv('GOOGLE_API_KEY')
if not API_KEY:
    raise ValueError('Please set the environment variable GOOGLE_API_KEY to your Gemini API key')
genai.configure(api_key=API_KEY)

MODEL_NAME = os.getenv('GEMINI_MODEL', 'gemini-pro')

# Translation function

def gemini_translate(text, target_language='en'):
    prompt = f"""
Only return the translated content, no explanation or formatting. Translate the following content to {target_language}:
{text}
"""
    model = genai.GenerativeModel(MODEL_NAME)
    response = model.generate_content(prompt)
    return response.text.strip()

def translate_excel(file_path):
    try:
        wb = openpyxl.load_workbook(file_path)
        print(f'[LOG] Source file opened: {file_path}')
    except Exception as e:
        print(f'[ERROR] Failed to open source file: {e}')
        return
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    print(f'[LOG] Headers: {headers}')
    print(f'[LOG] Total columns: {len(headers)}')
    print(f'[LOG] Total rows: {ws.max_row}')
    key_col = 1  # The first column is the key
    en_col = 2   # The second column is the English source
    lang_cols = list(range(3, len(headers) + 1))  # Target language columns
    lang_names = headers[2:]  # Target language names from header

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        key = row[key_col - 1].value
        en_text = row[en_col - 1].value
        print(f"[LOG] Row {row_idx} key: {key}, en_GB: {en_text}")
        if not en_text:
            continue
        for idx, col in enumerate(lang_cols):
            cell = row[col - 1]
            # Only translate if cell is None or blank (after stripping whitespace)
            if cell.value is None or str(cell.value).strip() == '':
                target_lang = lang_names[idx]
                retry_count = 0
                success = False
                while retry_count < 3 and not success:
                    try:
                        print(f"[LOG] Translating (row {row_idx}, key={key}, target={target_lang}, attempt {retry_count+1}): {en_text}")
                        translated = gemini_translate(en_text, target_lang)
                        cell.value = translated
                        print(f"[LOG] Translation done (row {row_idx}, key={key}, target={target_lang}): {translated}")
                        time.sleep(1)  # Avoid API rate limit
                        # Save after each cell translation
                        output_path = file_path.replace('.xlsx', '_translated.xlsx')
                        try:
                            wb.save(output_path)
                            print(f'[LOG] Saved after cell translation: {output_path}')
                        except Exception as e:
                            print(f'[ERROR] Failed to save after cell translation: {e}')
                        success = True
                    except Exception as e:
                        retry_count += 1
                        print(f"[ERROR] Translation failed (row {row_idx}, key={key}, target={target_lang}, attempt {retry_count}): {e}")
                        if retry_count >= 3:
                            print(f"[FATAL] Translation failed 3 times, exiting.")
                            output_path = file_path.replace('.xlsx', '_translated.xlsx')
                            try:
                                wb.save(output_path)
                                print(f'[LOG] File saved: {output_path}')
                            except Exception as e:
                                print(f'[ERROR] Failed to save file: {e}')
                            finally:
                                wb.close()
                            exit(1)
                print(f"[LOG] Current cell value (row {row_idx}, key={key}, lang={lang_names[idx]}): {cell.value}")
    output_path = file_path.replace('.xlsx', '_translated.xlsx')
    try:
        wb.save(output_path)
        print(f'[LOG] File saved: {output_path}')
    except Exception as e:
        print(f'[ERROR] Failed to save file: {e}')
    finally:
        wb.close()
    print(f'Translation finished. Output saved to: {output_path}')

if __name__ == '__main__':
    # Batch Excel translation
    excel_path = os.getenv('TRANSLATE_INPUT_FILE', 'solarreels_v1.xlsx')
    translate_excel(excel_path)
